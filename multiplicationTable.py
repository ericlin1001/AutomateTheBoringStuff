# python3
# multiplicationTable.py num - save multiplicationTable in a spreadsheet.
import openpyxl, openpyxl.cell
import os

#read command line.
if len(os.sys.argv) < 2:
    num = 5;
else:
    num = int(os.sys.argv[1]);

print("Creating multiplicationTable with num = %d, saving to m%d.xlsx"%(num,
    num));

#create the table.
wb = openpyxl.Workbook();
wb.remove_sheet(wb.active)
s = wb.create_sheet("multiplicationTable(num=%d)"%num);
table = [0] * num
for i in range(1, num + 1):
    table[i - 1] = [0] * num
    for j in range(1, num + 1):
        table[i - 1][j - 1] = i * j
        cn =openpyxl.cell.get_column_letter(j) + str(i) 
        s[cn] = table[i - 1][j - 1];

wb.save('m%d.xlsx'%num);





#save the table into spread sheet.
